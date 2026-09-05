"""浙江大学计算机科学与技术学院（www.cs.zju.edu.cn，真实站点在 /csen/ 路径下）。
- 教师名录 /csen/27003/list.htm：JS 渲染，条目链到 person.zju.edu.cn 统一主页（64 人）
- 导师资格目录文章页挂官方 Excel《各学科博导硕导-全》：
  4 学科（计算机/软件工程/网安/电子信息）× 博导/硕导 名单，按姓名对齐（418 人）
- 详情页 person.zju.edu.cn：姓名/职称/导师资格/单位/邮箱/地址/研究方向 服务端渲染，
  简介等内容在 column API（apiColumn 签名在页面 JS 变量里）
"""
import io
import re

from bs4 import BeautifulSoup

from crawler import fetch

_LIST = "http://www.cs.zju.edu.cn/csen/27003/list.htm"
_XLSX_ARTICLE = ("http://www.cs.zju.edu.cn/csen/2021/0915/"
                 "c27006a2421689/page.htm")
_PERSON = re.compile(r"person\.zju\.edu\.cn/([A-Za-z0-9_]+)/?")


def _norm(nm):
    return re.sub(r"[\s\u3000]+", "", nm)


def _load_xlsx_supervisors():
    """下载官方导师资格目录 xlsx，返回 {姓名: {"supervisor":..., "subjects":[...]}}。"""
    import openpyxl
    art, _, _ = fetch.fetch_text("GET", _XLSX_ARTICLE)
    m = re.search(r'href="([^"]+\.xlsx)"', art)
    if not m:
        return {}
    raw, _, _ = fetch.fetch("GET", "http://www.cs.zju.edu.cn" + m.group(1))
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    # 表头在前两行：每组首列是 "081200计算机科学与技术" 等学科名
    headers = []
    for r in wb.active.iter_rows(min_row=1, max_row=2, values_only=True):
        headers.append(list(r))
    disc_by_group = {}
    for gi in range(0, 15, 3):
        for r in headers:
            if gi < len(r) and r[gi] and re.match(r"^\d{6}", str(r[gi])):
                disc_by_group[gi] = re.sub(r"^\d{6}", "", str(r[gi]).strip())
    out = {}
    for row in wb.active.iter_rows(min_row=3, values_only=True):
        for gi in (0, 3, 6, 9):
            cells = row[gi:gi + 3] if len(row) >= gi + 3 else None
            if not cells:
                continue
            disc = disc_by_group.get(gi)
            for off, kind in ((1, "博导"), (2, "硕导")):
                v = cells[off] if len(cells) > off else None
                if not v:
                    continue
                nm = _norm(str(v))
                if not re.fullmatch(r"[\u4e00-\u9fa5·]{2,5}", nm):
                    continue
                rec = out.setdefault(nm, {"kinds": set(), "subjects": []})
                rec["kinds"].add(kind)
                if disc and disc not in rec["subjects"]:
                    rec["subjects"].append(disc)
    return {nm: {"supervisor": "、".join(sorted(r["kinds"])),
                 "subjects": r["subjects"]}
            for nm, r in out.items()}


def iter_roster(cfg):
    people = {}
    # 1) 教师名录（渲染页，条目链 person.zju.edu.cn）
    html, meta, _ = fetch.fetch_rendered(cfg["list"]["urls"][0], wait_ms=3000,
                                         scroll=True)
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.select("a[href]"):
        href = str(a.get("href") or "")
        if not _PERSON.search(href):
            continue
        nm = _norm(a.get_text(strip=True))
        u = href if href.startswith("http") else "http://www.cs.zju.edu.cn" + href
        if nm and u not in people:
            people[u] = {"name": nm, "url": u, "profile_url": u}
    # 2) 官方导师资格目录 xlsx：博导/硕导 + 学科，按姓名对齐
    sup = _load_xlsx_supervisors()
    for u, p in people.items():
        s = sup.get(p["name"])
        if s:
            p["supervisor"] = s["supervisor"]
            if s["subjects"]:
                p["subjects"] = "、".join(s["subjects"])
    # 3) xlsx 里有而名录没有的导师：入库为 roster 级（详情页落在目录文章页）
    known = {p["name"] for p in people.values()}
    for nm, s in sup.items():
        if nm in known:
            continue
        people[_XLSX_ARTICLE + "#" + nm] = {
            "name": nm, "url": _XLSX_ARTICLE + "#" + nm,
            "profile_url": _XLSX_ARTICLE,
            "supervisor": s["supervisor"],
            **({"subjects": "、".join(s["subjects"])} if s["subjects"] else {}),
        }
    return list(people.values()), meta


def parse_detail(cfg, html, url):
    if "person.zju.edu.cn" not in url:
        return {}  # xlsx-only 记录落在目录文章页，无可解析字段
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style"]):
        tag.decompose()
    out = {}
    # 头部区（到「个人简介」栏目为止）：姓名 博士 | 副教授 | 博士生导师 | 职务
    body_text = soup.get_text("\n", strip=True)
    head = body_text.split("个人简介", 1)[0][:500]
    tm = re.search(r"(讲席|特聘|长聘|兼职|客座|荣誉)?(教授|副教授|助理教授|讲师|"
                   r"研究员|副研究员|助理研究员|高级工程师|工程师|实验师|馆员)", head)
    if tm:
        out["title"] = tm.group(0)
    sup = sorted({("博导" if "博士" in s else "硕导")
                  for s in re.findall(r"(博士生导师|硕士生导师)", head)})
    if sup:
        out["supervisor"] = "、".join(sup)
    # personal_bottom 里的字段 li（模板差异较大，按 label 文本匹配）
    for li in soup.find_all("li"):
        t = li.get_text(" ", strip=True)
        if "email" not in out and t.startswith("邮箱"):
            m = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", t)
            if m:
                out["email"] = m.group(0)
        elif "office_address" not in out and t.startswith("地址"):
            v = t.split("地址", 1)[-1].strip(" ：:").strip()
            if v:
                out["office_address"] = v
        elif "phone" not in out and t.startswith("电话"):
            v = t.split("电话", 1)[-1].strip(" ：:").strip()
            if v:
                out["phone"] = v
        elif "institute_from_detail" not in out and t.startswith("单位"):
            v = t.split("单位", 1)[-1].strip(" ：:").strip()
            if v:
                out["institute_from_detail"] = [v]
    rd_items = [li.get_text(" ", strip=True).lstrip("· ").strip()
                for li in soup.select("ul.second_research li")]
    rd_items = [x for x in rd_items if x]
    if rd_items:
        out["research_directions"] = rd_items[:20]
    m = re.search(r"更新时间\s*[：:]\s*(\d{4}-\d{2}-\d{2})", body_text)
    if m:
        out["source_updated_at"] = m.group(1)
    # 个人简介等内容在 column API; 接口校验 session cookie(route/PHPSESSID),
    # 必须先用同一会话访问页面再调接口, 否则 Access denied
    apic = re.search(r"var apiColumn\s*=\s*[\"']([^\"']+)[\"']", html)
    puid = re.search(r"var pageUid\s*=\s*[\"']([^\"']+)[\"']", html)
    col_id = None
    nav = soup.select_one("#tab_nav")
    if nav:
        for li in nav.select("li"):
            if "个人简介" in li.get_text():
                col_id = li.get("col")
    if apic and puid and col_id:
        import json as _json
        import requests as _requests
        try:
            s = _requests.Session()
            s.headers["User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            r0 = s.get(url, timeout=20)  # 取 session cookie + 新鲜签名
            fresh_apic = re.search(r"var apiColumn\s*=\s*[\"']([^\"']+)[\"']", r0.text)
            if not fresh_apic:
                return out
            api = ("https://person.zju.edu.cn" + fresh_apic.group(1) +
                   f"&column_id={col_id}&pageUid={puid.group(1)}&type=1")
            r = s.get(api, headers={"Referer": url, "X-Requested-With": "XMLHttpRequest"},
                      timeout=20)
            data = _json.loads(r.text).get("data") or {}
            content = data.get("content") or ""
            bio = BeautifulSoup(content, "html.parser").get_text("\n", strip=True)
            bio = re.sub(r"\n{2,}", "\n", bio).strip()
            if len(bio) > 30:
                out["bio_raw"] = bio[:5000]
        except Exception:
            pass
    return out
